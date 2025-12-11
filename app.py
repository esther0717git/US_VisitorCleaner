import streamlit as st
import pandas as pd
import re
from io import BytesIO
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ───── Streamlit setup ────────────────────────────────────────────────────────
st.set_page_config(page_title="Visitor List Cleaner (US)", layout="wide")
st.title("🇺🇸 Clarity Gate - US Visitor Data Cleaning & Validation 🫧")

st.info(
    """
    **Data Integrity Is Our Foundation**  
    At every step—from file upload to final report—we enforce strict validation to guarantee your visitor data is accurate, complete, and compliant.  
    Maintaining integrity not only expedites gate clearance, it protects our facilities and ensures we meet all regulatory requirements.  

    **Why is Data Integrity Important?**  
    **Accuracy**: Correct visitor details reduce clearance delays.  
    **Security**: Reliable ID checks prevent unauthorized access.  
    **Compliance**: Audit-ready records ensure regulatory adherence.  
    **Efficiency**: Trustworthy data powers faster reporting and analytics.
    """
)

# ───── 3) Uploader & Warning ───────────────────────────────────────────────────
st.markdown(
    """<div style='font-size:16px; font-weight:bold; color:#38761d;'>
    Please ensure your spreadsheet has no missing or malformed fields.<br>
    Columns E (First Name) and Column F (Middle and Last Name) are not required to be filled in.<br>
    </div>""",
    unsafe_allow_html=True
)

# ───── Download Sample Template ───────────────────────────────────────────────
with open("US_Template.xlsx", "rb") as f:
    sample_bytes = f.read()
st.download_button(
    label="⬇️ Download US Template",
    data=sample_bytes,
    file_name="US_Template.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

# ───── 4) Estimate Clearance Date ───────────────────────────────────────────────
US_EASTERN = ZoneInfo("America/New_York")

def now_in_eastern() -> datetime:
    """Return current time converted from UTC to US/Eastern (handles DST correctly)."""
    return datetime.now(timezone.utc).astimezone(US_EASTERN)

now = now_in_eastern()
formatted_now = now.strftime("%A %d %B, %I:%M%p").lstrip("0")
st.write("**Today:**", formatted_now)

def next_working_day(d):
    """Return the next calendar date that is a weekday (Mon–Fri)."""
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d

def earliest_clearance_inclusive(submit_dt: datetime, workdays: int = 2) -> datetime:
    """
    Business rule:
    - Start counting working days from the submission calendar day itself (inclusive).
    - If submission falls on a weekend, start from the next Monday.
    - After counting N working days, the earliest clearance is the **next** working day.
    """
    # Day 1 is the submission date (roll forward if weekend)
    d = next_working_day(submit_dt.date())

    # Count remaining working days (we already counted Day 1)
    counted = 1
    while counted < workdays:
        d += timedelta(days=1)
        if d.weekday() < 5:
            counted += 1

    # Earliest clearance is the NEXT working day after the Nth working day
    clearance = d + timedelta(days=1)
    clearance = next_working_day(clearance)
    return clearance

if st.button("▶️ Earliest clearance:"):
    # Use Eastern 'now' to align with the displayed date
    clearance_date = earliest_clearance_inclusive(now, workdays=2)
    st.success(f" **{clearance_date:%A} {clearance_date.day} {clearance_date:%B}**")

# ───── Helper functions ────────────────────────────────────────────────────────
def split_name(full_name):
    s = str(full_name).strip()
    if " " in s:
        i = s.find(" ")
        return pd.Series([s[:i], s[i+1:]])
    return pd.Series([s, ""])

def clean_gender(g):
    v = str(g).strip().upper()
    if v == "M":
        return "Male"
    if v == "F":
        return "Female"
    if v in ("MALE", "FEMALE"):
        return v.title()
    return v.title()

def fix_mobile(x):
    d = re.sub(r"\D", "", str(x))
    # if too long...
    if len(d) > 10:
        extra = len(d) - 10
        if d.endswith("0" * extra):
            d = d[:-extra]
        else:
            d = d[-10:]
    if len(d) < 10:
        d = d.zfill(10)
    return d

# ───── Core Cleaning Logic ────────────────────────────────────────────────────
def clean_data_us(df: pd.DataFrame) -> pd.DataFrame:
    # 1) Trim to exactly 11 cols then rename
    df = df.iloc[:, :11]
    df.columns = [
        "S/N",
        "Vehicle Plate Number",
        "Company Full Name",
        "Full Name",
        "First Name",
        "Middle and Last Name",
        "Driver License Number",
        "Nationality (Country Name)",
        "Gender",
        "Mobile Number",
        "Remarks",
    ]

    # 2) Drop rows where all of Full Name → Mobile are blank
    df = df.dropna(subset=df.columns[3:10], how="all")

    # 3) Normalize nationality
    nat_map = {
        "chinese":     "China",
        "singaporean": "Singapore",
        "malaysian":   "Malaysia",
        "indian":      "India",
        "usa":         "United States",
        "us":          "United States",
    }
    df["Nationality (Country Name)"] = (
        df["Nationality (Country Name)"]
        .astype(str)
        .str.strip()
        .str.lower()
        .replace(nat_map, regex=False)
        .str.title()
    )

    # 4) Sort by Company → Country → Full Name
    df = df.sort_values(
        ["Company Full Name", "Nationality (Country Name)", "Full Name"],
        ignore_index=True,
    )

    # 5) Reset S/N
    df["S/N"] = range(1, len(df) + 1)

    # 6) Standardize Vehicle Plate Number
    df["Vehicle Plate Number"] = (
        df["Vehicle Plate Number"]
        .astype(str)
        .str.replace(r"[\/,]", ";", regex=True)
        .str.replace(r"\s*;\s*", ";", regex=True)
        .str.strip()
        .replace("nan", "", regex=False)
    )

    # 7) Proper-case & split names (in case the template didn't pre-split)
    df["Full Name"] = df["Full Name"].astype(str).str.title()
    df[["First Name", "Middle and Last Name"]] = df["Full Name"].apply(split_name)

    # 8) Fix Mobile Number → 10 digits
    df["Mobile Number"] = df["Mobile Number"].apply(fix_mobile)

    # 9) Normalize gender
    df["Gender"] = df["Gender"].apply(clean_gender)

    # 10) Driver License Number: remove spaces, keep last 4 digits
    df["Driver License Number"] = (
        df["Driver License Number"]
        .fillna("")                           # handle NaN
        .astype(str)
        .str.replace(r"\.0$", "", regex=True) # remove .0 endings
        .str.replace(r"\D", "", regex=True)   # keep only digits
        .str[-4:]                             # last 4 digits
    )

    return df

# ───── Build & style Excel while PRESERVING other sheets ─────────────────────
def generate_visitor_only_us(df: pd.DataFrame, uploaded_file) -> tuple[BytesIO, bool]:
    buf = BytesIO()
    has_errors = False  # track if any cell is highlighted red

    # Reload original workbook so all other sheets are preserved
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file)

    # Get or create Visitor List sheet
    if "Visitor List" in wb.sheetnames:
        ws = wb["Visitor List"]
        # Clear existing contents
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Visitor List")

    # Write DataFrame to Visitor List sheet
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # styling objects
    header_fill  = PatternFill("solid", fgColor="94B455")
    border       = Border(Side("thin"), Side("thin"), Side("thin"), Side("thin"))
    center       = Alignment("center", "center")
    normal_font  = Font(name="Calibri", size=9)
    bold_font    = Font(name="Calibri", size=9, bold=True)
    invalid_fill = PatternFill("solid", fgColor="E6B8B7")  # light red

    # 1) Apply borders, alignment, font
    for row in ws.iter_rows():
        for cell in row:
            cell.border    = border
            cell.alignment = center
            cell.font      = normal_font

    # 2) Style header row
    for col in range(1, ws.max_column + 1):
        h = ws[f"{get_column_letter(col)}1"]
        h.fill = header_fill
        h.font = bold_font

    # 3) Freeze top row
    ws.freeze_panes = "A2"

    # 4) Auto-fit columns & set row height
    for col in ws.columns:
        values = [len(str(cell.value)) for cell in col if cell.value is not None]
        width = max(values) if values else 10
        ws.column_dimensions[get_column_letter(col[0].column)].width = width + 2
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20

    # 5) Highlight invalid Driver License Number (Column G)
    dl_col_idx = 7  # Column G is the 7th column
    dl_pattern = re.compile(r"^\d{4}$")

    for r in range(2, ws.max_row + 1):  # skip header row
        cell = ws.cell(row=r, column=dl_col_idx)
        value = str(cell.value).strip() if cell.value is not None else ""
        if not dl_pattern.match(value):
            cell.fill = invalid_fill
            has_errors = True

    # 6) Highlight blank cells in Column F (Middle and Last Name)
    col_f_idx = 6  # Column F
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_f_idx)
        value = str(cell.value).strip() if cell.value not in (None, "") else ""
        if value == "":
            cell.fill = invalid_fill
            has_errors = True

    # 7) Vehicles summary
    plates = []
    for v in df["Vehicle Plate Number"].dropna():
        plates += [x.strip() for x in str(v).split(";") if x.strip()]
    ins = ws.max_row + 2
    if plates:
        ws[f"B{ins}"].value     = "Vehicles"
        ws[f"B{ins}"].border    = border
        ws[f"B{ins}"].alignment = center
        ws[f"B{ins+1}"].value   = ";".join(sorted(set(plates)))
        ws[f"B{ins+1}"].border  = border
        ws[f"B{ins+1}"].alignment = center
        ins += 2

    # 8) Total Visitors
    ws[f"B{ins}"].value     = "Total Visitors"
    ws[f"B{ins}"].border    = border
    ws[f"B{ins}"].alignment = center
    ws[f"B{ins+1}"].value   = df["Company Full Name"].notna().sum()
    ws[f"B{ins+1}"].border  = border
    ws[f"B{ins+1}"].alignment = center

    # Save full workbook (all sheets) to buffer
    wb.save(buf)
    buf.seek(0)
    return buf, has_errors

# ───── Streamlit UI: Upload & Download ────────────────────────────────────────
uploaded = st.file_uploader("📁 Upload File", type=["xlsx"])
if uploaded:
    raw_df = pd.read_excel(uploaded, sheet_name="Visitor List")
    cleaned = clean_data_us(raw_df)
    out_buf, has_errors = generate_visitor_only_us(cleaned, uploaded)

    # Status message based on validation
    if has_errors:
        st.error("❗ Error(s) found — please correct all red-highlighted cells before submission.")
    else:
        st.success("✅ Please double-check critical fields before sharing with DC team.")

    # Build filename: CompanyName_YYYYMMDD.xlsx using US/Eastern date (from UTC)
    today_eastern = now_in_eastern().strftime("%Y%m%d")
    company_cell = raw_df.iloc[0, 2]
    company = (
        str(company_cell).strip()
        if pd.notna(company_cell) and str(company_cell).strip()
        else "VisitorList"
    )
    fname = f"{company}_{today_eastern}.xlsx"

    st.download_button(
        label="📥 Download Cleaned Visitor List (US)",
        data=out_buf.getvalue(),
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ───── 5) Final Notice (always shown) ────────────────────────────────────────
st.markdown(
    """
    <div style="line-height:1.2; font-size:16px;">
      We will do our utmost to deliver your access ticket 1 day before your scheduled entry.<br>
      Kindly ensure that approved access clearance codes are obtained before planning or commencing any work activities in the data center.<br>
      Please be reminded to go through the Clarity Gate prior to submission, and ensure that all visitor and shipment details are complete and accurate to prevent rescheduling due to clarification.<br><br>
      <strong>Note:</strong><br>
      The Clarity Gate operates on the GOFAI system, which relies on explicitly programmed rules and logic.<br>
      Although its validation accuracy can reach up to 98%, we strongly recommend that you thoroughly review all information before submission.<br>
      Thank you for your cooperation.<br><br>
    </div>
    """,
    unsafe_allow_html=True,
)

# ───── 6) Vendor Accuracy Reminder (always shown) ────────────────────────────
st.markdown(
    """
    <div style="line-height:1.2; font-size:16px;">
      <strong>Kindly remind all vendors to take the accuracy of the submitted information seriously.</strong><br>
      Any <em>incorrect or incomplete details</em> will result in <em>rejection</em>, and the personnel will not be allowed to enter the data centre.<br>
      <em>This requirement is non-negotiable, and strict compliance is expected.</em><br>
      Please ensure this message is clearly conveyed to all concerned.
    </div>
    """,
    unsafe_allow_html=True,
)

